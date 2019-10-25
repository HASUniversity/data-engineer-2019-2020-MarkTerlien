from openpyxl import load_workbook

# Constant
INPUT_FILE = 'A16 Ri Antwerpen.xlsx'

# Open Excel file with micro rooster
excel_file = load_workbook(INPUT_FILE)

# Loop over sheets
for sheet in excel_file.worksheets:
    print(sheet)

# Get worksheet by index (0 = Overzicht, 1 = intentesiteit, 2 = Gemiddelde snelheid)
ws_index = 2
ws = excel_file[excel_file.sheetnames[ws_index]]
print(ws.title)

# get max row count
max_row=ws.max_row
print(max_row)

# get max column count
max_column=ws.max_column
print(max_column)

# iterate over all cells 
# iterate over all rows
for i in range(1,max_row+1):
     
     # iterate over all columns
     for j in range(1,max_column+1):
          
          # get particular cell value
          cell_obj=sheet.cell(row=i,column=j)
          
          # print cell value     
          print("Value for row " + str(i) + " and column " + str(j) + ": " + str(cell_obj.value))
